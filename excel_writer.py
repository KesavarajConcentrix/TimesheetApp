import calendar
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Light green for weekends
GREEN_FILL = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

# Yellow for empty weekdays
YELLOW_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

# Light blue header
HEADER_FILL = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")


# Border style - thin black border on all sides
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

def write_task_file(task_name, df, output_dir):
    """
    Generate Concentrix-style timesheet with borders:
    Row 1: Month title (e.g., Nov'25)
    Row 2: Weekday formulas
    Row 3: Headers + date columns
    Row 4+: Data rows per employee
    
    FIX: Calculate totals in Python, write values directly (not formulas)
    """
    import os
    os.makedirs(output_dir, exist_ok=True)

    # Detect month/year
    month_series = df["Reported Date"].dt.to_period("M")
    dominant_month = month_series.mode()[0]
    year = dominant_month.year
    month = dominant_month.month

    # Build full month date list
    last_day = calendar.monthrange(year, month)[1]
    first_date = date(year, month, 1)
    last_date = date(year, month, last_day)

    dates = []
    d = first_date
    while d <= last_date:
        dates.append(d)
        d += timedelta(days=1)

    # Aggregate per worker/day
    employee_names = sorted(df["Worker"].dropna().astype(str).unique().tolist())

    # Output filename
    safe_task = task_name.replace("/", "_")
    file_path = os.path.join(
        output_dir,
        f"Concentrix_{safe_task}-Timesheet-{month:02d}{year % 100:02d}.xlsx"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet"

    # Row 1: Title (center, bold, bigger font)
    ws["A1"] = f"{calendar.month_abbr[month]}'{year % 100:02d}"
    ws["A1"].font = Font(size=14, bold=True)  # slightly bigger than other text
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:AJ1")


    # Row 2: Weekday formulas (start from column G)
    for col_idx in range(7, 7 + len(dates)):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=2, column=col_idx)
        cell.value = f'=TEXT(WEEKDAY({col_letter}3,1),"ddd")'
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row 3: Headers
    headers = ["Region", "Emp Name", "Total Hours", "Total days", "Manager", ""]
    for col_idx, header in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col_idx)
        c.value = header
        c.fill = HEADER_FILL
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER

    # Row 3: Date headers
    for i, dt in enumerate(dates):
        col_idx = 7 + i
        c = ws.cell(row=3, column=col_idx)
        c.value = dt
        c.number_format = "DD-MMM-YY"
        c.fill = HEADER_FILL
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER

    # Data rows start at row 4
    row_idx = 4
    for emp in employee_names:
        # Region
        c_region = ws.cell(row=row_idx, column=1)
        c_region.value = task_name
        c_region.border = THIN_BORDER
        c_region.alignment = Alignment(horizontal="left", vertical="center")

        # Employee Name
        c_emp = ws.cell(row=row_idx, column=2)
        c_emp.value = emp
        c_emp.border = THIN_BORDER
        c_emp.alignment = Alignment(horizontal="left", vertical="center")

        # Filter data for this employee
        emp_df = df[df["Worker"].astype(str) == emp]

        # Calculate daily hours (FIX: Do this in Python)
        daily_hours = {}
        for dt in dates:
            daily_hours[dt] = emp_df.loc[emp_df["Reported Date"].dt.date == dt, "Hours"].sum()

        # Total Hours: Sum all hours for this employee (FIX: Write value, not formula)
        total_hours = sum(daily_hours.values())
        c_total_hrs = ws.cell(row=row_idx, column=3)
        c_total_hrs.value = int(total_hours) if total_hours == int(total_hours) else total_hours
        c_total_hrs.border = THIN_BORDER
        c_total_hrs.alignment = Alignment(horizontal="center", vertical="center")
        c_total_hrs.number_format = "0"

        # Determine dynamic hours-per-day divisor
        # Default to 8; override if 8 or 9 is actually used in daily hours
        hours_per_day = 8.0
        if any(h == 8 for h in daily_hours.values()):
            hours_per_day = 8.0
        elif any(h == 9 for h in daily_hours.values()):
            hours_per_day = 9.0

        # Total Days: Total Hours / dynamic hours_per_day (write value, not formula)
        total_days = total_hours / hours_per_day if hours_per_day > 0 else 0
        c_total_days = ws.cell(row=row_idx, column=4)
        c_total_days.value = round(total_days, 1)
        c_total_days.border = THIN_BORDER
        c_total_days.alignment = Alignment(horizontal="center", vertical="center")
        c_total_days.number_format = "0.0"

        # Manager (blank)
        c_mgr = ws.cell(row=row_idx, column=5)
        c_mgr.value = ""
        c_mgr.border = THIN_BORDER
        c_mgr.alignment = Alignment(horizontal="left", vertical="center")

        # Separator (blank)
        c_sep = ws.cell(row=row_idx, column=6)
        c_sep.value = ""
        c_sep.border = THIN_BORDER

        # Fill daily hours
        for i, dt in enumerate(dates):
            col_idx = 7 + i
            cell = ws.cell(row=row_idx, column=col_idx)

            hours = daily_hours[dt]
            if hours > 0:
                cell.value = int(hours) if hours == int(hours) else float(hours)

            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = "0"
            cell.border = THIN_BORDER

            # Coloring
            if dt.weekday() >= 5:  # Weekend
                cell.fill = GREEN_FILL
            else:  # Weekday
                if hours <= 0:
                    cell.fill = YELLOW_FILL

        row_idx += 1

    # Legend
    legend_row = row_idx + 2

    c_green = ws.cell(row=legend_row, column=1)
    c_green.fill = GREEN_FILL
    c_green.border = THIN_BORDER

    c_green_text = ws.cell(row=legend_row, column=2)
    c_green_text.value = "Sat / Sun"
    c_green_text.border = THIN_BORDER

    c_yellow = ws.cell(row=legend_row + 1, column=1)
    c_yellow.fill = YELLOW_FILL
    c_yellow.border = THIN_BORDER

    c_yellow_text = ws.cell(row=legend_row + 1, column=2)
    c_yellow_text.value = "Empty weekday"
    c_yellow_text.border = THIN_BORDER

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 5
    for col_idx in range(7, 7 + len(dates)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 10

    wb.save(file_path)
    print(f"✅ Generated: {file_path}")
    return file_path
